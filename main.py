					# Was passiert hier?
					# Wir laden alle Bibliotheken die wir brauchen. re für Textsuche, openpyxl für Excel, datetime für Datumsverarbeitung, 
					# scipy für die Korrelationsberechnung. Falls scipy nicht installiert ist: pip install scipy

import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from datetime import datetime
from scipy import stats

					# Was passiert hier?
					# Diese Hilfsfunktion wird überall verwendet. Sie sucht nach einem Muster im Text und gibt den Treffer zurück. Findet sie 
					# nichts (z.B. weil ein altes Log den neuen Abschnitt nicht hat), gibt sie einfach einen leeren String zurück – kein Absturz.

def re_extract(pattern, text, fallback=""):
    treffer = re.search(pattern, text)
    return treffer.group(1).strip() if treffer else fallback

					# Was passiert hier?
					# Samsung schreibt Zeiten als "7 Std 23 Min". Excel kann damit nicht rechnen. Diese Funktion wandelt das in eine einzige Zahl 
					# um (z.B. 443 Minuten), mit der Excel dann Diagramme und Berechnungen machen kann.

def zeit_zu_minuten(zeitstring):
    if not zeitstring or zeitstring == "---":
        return None
    std = re.search(r"(\d+)\s*Std", zeitstring)
    min_ = re.search(r"(\d+)\s*Min", zeitstring)
    stunden = int(std.group(1)) if std else 0
    minuten = int(min_.group(1)) if min_ else 0
    return stunden * 60 + minuten

					# Was passiert hier?
					# Die Wassereinträge sehen so aus: "2x 500ml = 1000ml". Wir wollen nur die Zahl in ml haben. Diese Funktion sucht alle ml-
					# Angaben und nimmt die letzte – das ist meistens die berechnete Gesamtsumme.

def wasser_ml_extrahieren(text):
    if not text or text.lower() == "kein":
        return None
    treffer = re.findall(r"(\d+)\s*ml", text)
    return int(treffer[-1]) if treffer else None

					# Was passiert hier?
					# Die .md Datei wird komplett eingelesen. Dann wird sie an den ====================== Trennlinien aufgeteilt. Jeder 
					# Block der "ENERGIEWERT LOG" enthält wird als einzelner Eintrag behandelt.

def parse_log(md_datei):
    with open(md_datei, "r", encoding="utf-8") as f:
        inhalt = f.read()

    eintraege = re.split(r"======================", inhalt)
    eintraege = [e.strip() for e in eintraege if "ENERGIEWERT LOG" in e]

    alle_daten = []

					# Was passiert hier?
					# Datum, Wochentag und Energiewert werden ausgelesen. Das Datum wird zusätzlich als echtes Datumsobjekt gespeichert 
					# ("Datum Sort") damit die Tabelle später korrekt nach Zeit sortiert werden kann.

    for eintrag in eintraege:
        daten = {}

        # === Basisdaten ===
        datum_raw = re_extract(r"Datum:\s*(\d{2}\.\d{2}\.\d{4})", eintrag)
        daten["Datum"] = datum_raw
        try:
            daten["Datum (Sort)"] = datetime.strptime(datum_raw, "%d.%m.%Y")
        except:
            daten["Datum (Sort)"] = None

        daten["Wochentag"] = re_extract(r"Datum:.*?\((.+?)\)", eintrag)
        energie_raw = re_extract(r"Energiewert:\s*(\d+)", eintrag)
        daten["Energiewert"] = int(energie_raw) if energie_raw else None
        daten["Energiestatus"] = re_extract(r'Energiewert:.*?"(.+?)"', eintrag)

					# Was passiert hier?
					# Alle Schlafdaten werden ausgelesen. Die Schlafdauer wird zusätzlich in Minuten umgerechnet (z.B. "7 Std 23 Min" → 443) 
					# damit Excel damit rechnen kann.

        # === Schlaf ===
        daten["Schlafbeginn"]      = re_extract(r"Schlafbeginn:\s*(.+?)\s*Uhr", eintrag)
        daten["Aufwachzeit"]       = re_extract(r"Aufwachzeit:\s*(.+?)\s*Uhr", eintrag)
        daten["Schlafdauer"]       = re_extract(r"Schlafdauer gesamt:\s*(.+)", eintrag)
        daten["Schlafdauer (Min)"] = zeit_zu_minuten(daten["Schlafdauer"])
        daten["Schlafpuls (bpm)"]  = re_extract(r"Schlafpuls:\s*(\d+)\s*bpm", eintrag)
        daten["HRV (ms)"]          = re_extract(r"HRV:\s*(\d+)\s*ms", eintrag)
        daten["Atmung (/min)"]     = re_extract(r"Atmung:\s*([\d,]+)\s*/min", eintrag)
        daten["Hauttemperatur"]    = re_extract(r"Hauttemperatur:\s*(.+)", eintrag)

					# Was passiert hier?
					# Jede Schlafphase wird zweimal gespeichert: einmal als lesbarer Text ("1 Std 34 Min") und einmal als reine Minutenzahl für 
					# Diagramme und Berechnungen.

        # === Schlafphasen ===
        rem_raw   = re_extract(r"REM-Schlaf:\s*(.+)", eintrag)
        licht_raw = re_extract(r"Leichtschlaf:\s*(.+)", eintrag)
        tief_raw  = re_extract(r"Tiefschlaf:\s*(.+)", eintrag)
        wach_raw  = re_extract(r"Wachphasen:\s*(.+)", eintrag)

        daten["REM-Schlaf"]          = rem_raw
        daten["REM (Min)"]           = zeit_zu_minuten(rem_raw)
        daten["Leichtschlaf"]        = licht_raw
        daten["Leichtschlaf (Min)"]  = zeit_zu_minuten(licht_raw)
        daten["Tiefschlaf"]          = tief_raw
        daten["Tiefschlaf (Min)"]    = zeit_zu_minuten(tief_raw)
        daten["Wachphasen"]          = wach_raw
        daten["Wachphasen (Min)"]    = zeit_zu_minuten(wach_raw)

					# Was passiert hier?
					# Schritte, aktive Zeit und Art der Aktivität werden ausgelesen. "--- Min" (kein Wert) wird von zeit_zu_minuten automatisch 
					# als None behandelt.

        # === Aktivität ===
        aktiv_raw = re_extract(r"Aktive Zeit:\s*(.+)", eintrag)
        daten["Schritte"]          = re_extract(r"Schritte:\s*(\d+)", eintrag)
        daten["Aktive Zeit"]       = aktiv_raw
        daten["Aktive Zeit (Min)"] = zeit_zu_minuten(aktiv_raw)
        daten["Aktivität Art"]     = re_extract(r"Art der Arbeit/Aktivität:\s*\[(.+?)\]", eintrag)

					# Was passiert hier?
					# Konsum wird als Text gespeichert. Zusätzlich gibt es eine einfache ja/nein Spalte für Cannabis (damit die Korrelation später 
					# funktioniert) und die Anzahl der Holy Energy Portionen als Zahl.

        # === Konsum ===
        daten["Holy Energy"]    = re_extract(r"Holy Energy:\s*\[(.+?)\]", eintrag)
        daten["Kaffee/Koffein"] = re_extract(r"Kaffee/sonstiges Koffein:\s*\[(.+?)\]", eintrag)
        daten["Cannabis"]       = re_extract(r"Cannabis:\s*\[(.+?)\]", eintrag)

        # Hilfsspalten für Auswertung: Wurde konsumiert? (ja/nein)
        daten["Cannabis (ja/nein)"] = "nein" if daten["Cannabis"].lower() == "kein" else "ja"
        daten["Holy (Anzahl)"] = re_extract(r"(\d+)x", daten["Holy Energy"])

					# Was passiert hier?
					# Die neue Wasser-Sektion wird ausgelesen. Da alte Logs diese Sektion nicht haben, gibt re_extract dort einfach einen 
					# leeren String zurück und wasser_ml_extrahieren gibt dann None zurück → leere Zelle in Excel. Kein Absturz.

        # === Wasser ===
        wasser_abschnitt     = re.search(r"--- Wasser Vortag ---(.*?)---", eintrag, re.DOTALL)
        wasser_text          = wasser_abschnitt.group(1) if wasser_abschnitt else ""

        wasser_holy_raw      = re_extract(r"Holy Energy:\s*\[(.+?)\]", wasser_text)
        wasser_hydration_raw = re_extract(r"Holy Hydration:\s*\[(.+?)\]", wasser_text)
        wasser_kaffee_raw    = re_extract(r"Kaffee:\s*\[(.+?)\]", wasser_text)
        wasser_tee_raw       = re_extract(r"Tee:\s*\[(.+?)\]", wasser_text)
        wasser_sonstiges_raw = re_extract(r"Sonstiges:\s*\[(.+?)\]", wasser_text)
        wasser_gesamt_raw    = re_extract(r"Gesamt:\s*\[(.+?)\]", wasser_text)

        daten["Wasser: Holy Energy"]   = wasser_ml_extrahieren(wasser_holy_raw)
        daten["Wasser: Hydration"]     = wasser_ml_extrahieren(wasser_hydration_raw)
        daten["Wasser: Kaffee"]        = wasser_ml_extrahieren(wasser_kaffee_raw)
        daten["Wasser: Tee"]           = wasser_ml_extrahieren(wasser_tee_raw)
        daten["Wasser: Sonstiges"]     = wasser_ml_extrahieren(wasser_sonstiges_raw)
        daten["Wasser: Gesamt (ml)"]   = wasser_ml_extrahieren(wasser_gesamt_raw)

					# Was passiert hier?
					# Körperdaten werden ausgelesen. An Tagen ohne Körper-Sektion (also alle außer Montag) gibt re_extract einfach einen 
					# leeren String zurück → leere Zelle in Excel.

        # === Körper (nur montags) ===
        daten["Gewicht (kg)"]      = re_extract(r"Gewicht:\s*([\d,]+)\s*kg", eintrag)
        daten["Körperfett (%)"]    = re_extract(r"Körperfett:\s*([\d,]+)\s*%", eintrag)
        daten["Muskelmasse (kg)"]  = re_extract(r"Muskelmasse:\s*([\d,]+)\s*kg", eintrag)
        daten["Körperwasser (%)"]  = re_extract(r"Körperwasser:\s*([\d,]+)\s*%", eintrag)
        daten["Skelettmasse (kg)"] = re_extract(r"Skelettmasse:\s*([\d,]+)\s*kg", eintrag)

					# Was passiert hier?
					# Die Samsung Bewertungen (Ausgezeichnet/Gut/Ausreichend/Achtung) werden ausgelesen.

        # === Samsung Faktoren ===
        daten["Ø Schlafzeit"]          = re_extract(r"Durchschnitt Schlafzeit:\s*\[(.+?)\]", eintrag)
        daten["Regelmäßigkeit"]        = re_extract(r"Regelmäßigkeit Schlafzeit:\s*\[(.+?)\]", eintrag)
        daten["Schlafregelmäßigkeit"]  = re_extract(r"Schlafregelmäßigkeit:\s*\[(.+?)\]", eintrag)
        daten["Schlafzeit"]            = re_extract(r"Schlafzeit:\s*\[(.+?)\]", eintrag)
        daten["Aktivität Vortag"]      = re_extract(r"Aktivität Vortag:\s*\[(.+?)\]", eintrag)
        daten["Aktivitätskontinuität"] = re_extract(r"Aktivitätskontinuität:\s*\[(.+?)\]", eintrag)
        daten["Samsung Schlafpuls"]    = re_extract(r"Schlafpuls:\s*\[(.+?)\]", eintrag)
        daten["Samsung HRV"]           = re_extract(r"Schlaf-HRV:\s*\[(.+?)\]", eintrag)

					# Was passiert hier?
					# Stress ist jetzt in zwei Felder aufgeteilt: Watch-Messung und subjektives Empfinden. Alte Logs haben nur ein Stress-Feld – 
					# das landet dann in "Stress (Watch)", "Stress (subjektiv)" bleibt leer.

        # === Kontext ===
        daten["Stress (Watch)"]      = re_extract(r"Stress \(Watch\):\s*\[(.+?)\]", eintrag)
        daten["Stress (subjektiv)"]  = re_extract(r"Stress \(subjektiv\):\s*\[(.+?)\]", eintrag)
        daten["Besonderheiten"]      = re_extract(r"Besonderheiten:\s*\[(.+?)\]", eintrag)

					# Was passiert hier?
					# Der fertig geparste Tag wird zur Liste hinzugefügt. Am Ende wird alles nach Datum sortiert, damit die Tabelle 
					# chronologisch ist.

        alle_daten.append(daten)

    alle_daten.sort(key=lambda x: x["Datum (Sort)"] or datetime.min)
    return alle_daten

					# Was passiert hier?
					# Das Workbook (die Excel-Datei) wird erstellt. Die Catppuccin Mocha Farben werden als Konstanten definiert damit wir sie 
					# überall wiederverwenden können ohne sie jedes Mal neu eintippen zu müssen.

def in_excel_exportieren(alle_daten, ausgabe_datei):
    wb = openpyxl.Workbook()

    # Farben (Catppuccin Mocha)
    FARBE_HEADER     = "1e1e2e"  # Dunkelblau-Lila (Hintergrund Kopfzeile)
    FARBE_HEADER_2   = "313244"  # Etwas heller (zweite Kopfzeile)
    FARBE_ZELLE      = "2a2a3e"  # Karten-Hintergrund
    FARBE_TEXT       = "cdd6f4"  # Hellgrau (Text)
    FARBE_GRUEN      = "a6e3a1"  # Mint-Grün (gute Werte)
    FARBE_ROT        = "f38ba8"  # Rosa-Rot (schlechte Werte)
    FARBE_BLAU       = "89b4fa"  # Hellblau (neutrale Highlights)

					# Was passiert hier?
					# Wir definieren welche Spalten in welcher Reihenfolge in der Excel-Tabelle erscheinen sollen. Spalten die hier nicht stehen 
					# (z.B. "Datum (Sort)") werden nicht angezeigt – die sind nur intern für die Sortierung.

    # =====================
    # TAB 1: ROHDATEN
    # =====================
    ws1 = wb.active
    ws1.title = "Rohdaten"
    ws1.sheet_view.showGridLines = False

    spalten = [
        "Datum", "Wochentag", "Energiewert", "Energiestatus",
        "Schlafbeginn", "Aufwachzeit", "Schlafdauer", "Schlafdauer (Min)",
        "Schlafpuls (bpm)", "HRV (ms)", "Atmung (/min)", "Hauttemperatur",
        "Wachphasen", "Wachphasen (Min)", "REM-Schlaf", "REM (Min)",
        "Leichtschlaf", "Leichtschlaf (Min)", "Tiefschlaf", "Tiefschlaf (Min)",
        "Schritte", "Aktive Zeit", "Aktive Zeit (Min)", "Aktivität Art",
        "Holy Energy", "Holy (Anzahl)", "Kaffee/Koffein", "Cannabis", "Cannabis (ja/nein)",
        "Wasser: Holy Energy", "Wasser: Hydration", "Wasser: Kaffee",
        "Wasser: Tee", "Wasser: Sonstiges", "Wasser: Gesamt (ml)",
        "Gewicht (kg)", "Körperfett (%)", "Muskelmasse (kg)",
        "Körperwasser (%)", "Skelettmasse (kg)",
        "Ø Schlafzeit", "Regelmäßigkeit", "Schlafregelmäßigkeit", "Schlafzeit",
        "Aktivität Vortag", "Aktivitätskontinuität", "Samsung Schlafpuls", "Samsung HRV",
        "Stress (Watch)", "Stress (subjektiv)", "Besonderheiten"
    ]

					# Was passiert hier?
					# Die erste Zeile wird als Kopfzeile geschrieben. Jede Spalte bekommt einen fett geschriebenen Titel auf dem dunklen 
					# Catppuccin-Hintergrund. Die Zeile wird etwas höher gemacht damit lange Spaltennamen lesbar sind.

    # Kopfzeile schreiben
    header_font   = Font(bold=True, color=FARBE_TEXT)
    header_fill   = PatternFill("solid", fgColor=FARBE_HEADER)
    header_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, spalte in enumerate(spalten, start=1):
        zelle = ws1.cell(row=1, column=col_idx, value=spalte)
        zelle.font      = header_font
        zelle.fill      = header_fill
        zelle.alignment = header_align

    ws1.row_dimensions[1].height = 40

					# Was passiert hier?
					# Jeder Tag (jedes Dictionary aus der Liste) wird als eine Zeile geschrieben. Wenn ein Wert nicht existiert (z.B. Körperdaten 
					# an einem Nicht-Montag), wird einfach eine leere Zelle geschrieben.

    # Daten schreiben
    zelle_fill = PatternFill("solid", fgColor=FARBE_ZELLE)
    zelle_font = Font(color=FARBE_TEXT)

    for row_idx, tag in enumerate(alle_daten, start=2):
        for col_idx, spalte in enumerate(spalten, start=1):
            wert = tag.get(spalte, "")
            zelle = ws1.cell(row=row_idx, column=col_idx, value=wert)
            zelle.fill      = zelle_fill
            zelle.font      = zelle_font
            zelle.alignment = Alignment(horizontal="left", vertical="center")

					# Was passiert hier?
					# Jede Spalte bekommt eine sinnvolle Breite. "Besonderheiten" ist z.B. sehr breit weil dort langer Text steht, "HRV (ms)" ist 
					# schmal weil dort nur 2-3 Ziffern stehen. freeze_panes = "A2" fixiert die Kopfzeile – beim Scrollen bleibt sie immer 
					# sichtbar.

    # Spaltenbreiten setzen
    spalten_breiten = {
        "Datum": 14, "Wochentag": 12, "Energiewert": 13, "Energiestatus": 22,
        "Schlafbeginn": 13, "Aufwachzeit": 13, "Schlafdauer": 18, "Schlafdauer (Min)": 16,
        "Schlafpuls (bpm)": 15, "HRV (ms)": 10, "Atmung (/min)": 13, "Hauttemperatur": 22,
        "Wachphasen": 15, "Wachphasen (Min)": 16, "REM-Schlaf": 15, "REM (Min)": 12,
        "Leichtschlaf": 15, "Leichtschlaf (Min)": 18, "Tiefschlaf": 15, "Tiefschlaf (Min)": 16,
        "Schritte": 10, "Aktive Zeit": 14, "Aktive Zeit (Min)": 16, "Aktivität Art": 35,
        "Holy Energy": 30, "Holy (Anzahl)": 14, "Kaffee/Koffein": 30,
        "Cannabis": 20, "Cannabis (ja/nein)": 16,
        "Wasser: Holy Energy": 20, "Wasser: Hydration": 18, "Wasser: Kaffee": 15,
        "Wasser: Tee": 12, "Wasser: Sonstiges": 16, "Wasser: Gesamt (ml)": 18,
        "Gewicht (kg)": 13, "Körperfett (%)": 14, "Muskelmasse (kg)": 16,
        "Körperwasser (%)": 16, "Skelettmasse (kg)": 16,
        "Ø Schlafzeit": 14, "Regelmäßigkeit": 16, "Schlafregelmäßigkeit": 20,
        "Schlafzeit": 13, "Aktivität Vortag": 16, "Aktivitätskontinuität": 20,
        "Samsung Schlafpuls": 18, "Samsung HRV": 13,
        "Stress (Watch)": 35, "Stress (subjektiv)": 18, "Besonderheiten": 60
    }

    for col_idx, spalte in enumerate(spalten, start=1):
        breite = spalten_breiten.get(spalte, 15)
        ws1.column_dimensions[get_column_letter(col_idx)].width = breite

    ws1.freeze_panes = "A2"

					# Was passiert hier?
					# Tab 2 bekommt eine kleine Hilfstabelle mit nur den Werten die für Diagramme gebraucht werden. Excel-Diagramme 
					# müssen ihre Daten aus einer sichtbaren Tabelle ziehen – deshalb steht sie hier. Die volle Rohdatentabelle ist in Tab 1.

    # =====================
    # TAB 2: ÜBERSICHT
    # =====================
    ws2 = wb.create_sheet("Übersicht")
    ws2.sheet_view.showGridLines = False

    # Hilfstabelle für Diagramme
    hilfs_spalten = ["Datum", "Energiewert", "HRV (ms)", "Schlafdauer (Min)",
                     "REM (Min)", "Tiefschlaf (Min)", "Leichtschlaf (Min)", "Schritte"]

    # Kopfzeile der Hilfstabelle
    for col_idx, spalte in enumerate(hilfs_spalten, start=1):
        zelle = ws2.cell(row=1, column=col_idx, value=spalte)
        zelle.font = Font(bold=True, color=FARBE_TEXT)
        zelle.fill = PatternFill("solid", fgColor=FARBE_HEADER)

    # Daten in die Hilfstabelle schreiben – explizit als Zahlen konvertieren
    for row_idx, tag in enumerate(alle_daten, start=2):
        for col_idx, spalte in enumerate(hilfs_spalten, start=1):
            wert = tag.get(spalte, "")
            # Strings in Zahlen umwandeln wo möglich
            if spalte != "Datum" and wert not in (None, ""):
                try:
                    wert = float(str(wert).replace(",", "."))
                    if wert == int(wert):
                        wert = int(wert)
                except (ValueError, TypeError):
                    wert = ""
            zelle = ws2.cell(row=row_idx, column=col_idx, value=wert)
            zelle.fill = PatternFill("solid", fgColor=FARBE_ZELLE)
            zelle.font = Font(color=FARBE_TEXT)

    anzahl_zeilen = len(alle_daten) + 1

					# Was passiert hier?
					# Ein Liniendiagramm wird erstellt das den Energiewert über alle Tage zeigt. Reference sagt dem Diagramm wo seine 
					# Daten stehen (Spalte 2 in Tab 2, von Zeile 1 bis zur letzten Zeile). Das Diagramm wird bei Zelle J2 platziert – also rechts 
					# neben der Hilfstabelle.

    # --- Diagramm 1: Energiewert über Zeit ---
    chart1 = LineChart()
    chart1.title   = "Energiewert über Zeit"
    chart1.style   = 10
    chart1.y_axis.title = "Energiewert"
    chart1.x_axis.title = "Tag"
    chart1.width   = 25
    chart1.height  = 15

    energie_daten = Reference(ws2, min_col=2, max_col=2, min_row=1, max_row=anzahl_zeilen)
    chart1.add_data(energie_daten, titles_from_data=True)
    ws2.add_chart(chart1, "J2")

					# Was passiert hier?
					# Ein gestapeltes Balkendiagramm zeigt REM, Tiefschlaf und Leichtschlaf pro Nacht übereinander. So siehst du sofort wie 
					# sich deine Schlafzusammensetzung über Zeit verändert und wie lang du insgesamt geschlafen hast.

    # --- Diagramm 2: Schlafphasen gestapelt ---
    chart2 = BarChart()
    chart2.type    = "bar"
    chart2.grouping = "stacked"
    chart2.title   = "Schlafphasen pro Nacht"
    chart2.y_axis.title = "Minuten"
    chart2.width   = 25
    chart2.height  = 15

    for col_idx, name in [(5, "REM"), (6, "Tiefschlaf"), (7, "Leichtschlaf")]:
        daten_ref = Reference(ws2, min_col=col_idx, max_col=col_idx,
                              min_row=1, max_row=anzahl_zeilen)
        chart2.add_data(daten_ref, titles_from_data=True)

    ws2.add_chart(chart2, "J35")

					# Was passiert hier?
					# HRV als Linie (Trend gut sichtbar) und Schritte als Balken (Tagesvergleich gut sichtbar). Beide werden unterhalb der ersten 
					# zwei Diagramme platziert.

    # --- Diagramm 3: HRV über Zeit ---
    chart3 = LineChart()
    chart3.title   = "HRV über Zeit"
    chart3.y_axis.title = "HRV (ms)"
    chart3.width   = 25
    chart3.height  = 15

    hrv_daten = Reference(ws2, min_col=3, max_col=3, min_row=1, max_row=anzahl_zeilen)
    chart3.add_data(hrv_daten, titles_from_data=True)
    ws2.add_chart(chart3, "J68")

    # --- Diagramm 4: Schritte über Zeit ---
    chart4 = BarChart()
    chart4.title   = "Schritte pro Tag"
    chart4.y_axis.title = "Schritte"
    chart4.width   = 25
    chart4.height  = 15

    schritte_daten = Reference(ws2, min_col=8, max_col=8, min_row=1, max_row=anzahl_zeilen)
    chart4.add_data(schritte_daten, titles_from_data=True)
    ws2.add_chart(chart4, "J101")

					# Was passiert hier?
					# Oben in Tab 3 steht jetzt ein vollständiger Erklärungstext der genau erklärt was Korrelation bedeutet und wie man die 
					# Werte liest. Fett markierte Zeilen sind Überschriften. Darunter kommt dann die eigentliche Korrelationstabelle.

    # =====================
    # TAB 3: KORRELATIONEN
    # =====================
    ws3 = wb.create_sheet("Korrelationen")
    ws3.sheet_view.showGridLines = False

    # --- Erklärungstext oben ---
    erklaerung = [
        ("Was zeigt diese Tabelle?", True),
        ("", False),
        ("Diese Tabelle zeigt wie stark verschiedene Faktoren mit deinem Energiewert zusammenhängen.", False),
        ("", False),
        ("Der 'Korrelationswert' ist eine Zahl zwischen -1 und +1:", True),
        ("", False),
        ("  +1.0  →  Perfekter positiver Zusammenhang", False),
        ("          Beispiel: Immer wenn du länger schläfst, ist dein Energiewert höher.", False),
        ("", False),
        ("  +0.5  →  Schwacher positiver Zusammenhang", False),
        ("          Tendenz vorhanden, aber nicht immer.", False),
        ("", False),
        ("   0.0  →  Kein Zusammenhang", False),
        ("          Dieser Faktor beeinflusst deinen Energiewert nicht.", False),
        ("", False),
        ("  -0.5  →  Schwacher negativer Zusammenhang", False),
        ("          Tendenz: Mehr davon = weniger Energie.", False),
        ("", False),
        ("  -1.0  →  Perfekter negativer Zusammenhang", False),
        ("          Immer wenn dieser Wert steigt, sinkt dein Energiewert.", False),
        ("", False),
        ("Faustregel:", True),
        ("  Über +0.5 oder unter -0.5  →  Interessant, da ist was dran!", False),
        ("  Zwischen -0.3 und +0.3    →  Eher zufällig, nicht überinterpretieren.", False),
        ("", False),
        ("Wichtig: Korrelation bedeutet nicht Ursache!", True),
        ("  Beispiel: Hohe Schritte könnten mit niedrigem Energiewert korrelieren,", False),
        ("  weil du an langen Arbeitstagen viel läufst UND erschöpft bist –", False),
        ("  nicht weil das Laufen selbst dich müde macht.", False),
        ("", False),
        ("─" * 60, False),
    ]

    for zeile_idx, (text, fett) in enumerate(erklaerung, start=1):
        zelle = ws3.cell(row=zeile_idx, column=1, value=text)
        zelle.font = Font(bold=fett, color=FARBE_TEXT, size=11)
        zelle.fill = PatternFill("solid", fgColor=FARBE_HEADER)

    ws3.column_dimensions["A"].width = 70
    ws3.column_dimensions["B"].width = 20
    ws3.column_dimensions["C"].width = 25
    ws3.column_dimensions["D"].width = 40

    start_zeile = len(erklaerung) + 2

					# Was passiert hier?
					# Die Korrelationstabelle bekommt Spaltenüberschriften. Dann wird definiert welche Faktoren mit dem Energiewert 
					# verglichen werden sollen. Die zweite Spalte im Tupel ist die Frage die dahinter steckt – die wird in der "Bedeutung" Spalte 
					# angezeigt.

    # --- Korrelationstabelle ---
    korr_header = ["Faktor", "Korrelationswert", "Stärke", "Bedeutung"]
    for col_idx, titel in enumerate(korr_header, start=1):
        zelle = ws3.cell(row=start_zeile, column=col_idx, value=titel)
        zelle.font = Font(bold=True, color=FARBE_TEXT)
        zelle.fill = PatternFill("solid", fgColor=FARBE_HEADER_2)

    # Welche Faktoren sollen mit dem Energiewert verglichen werden?
    faktoren = [
        ("Schlafdauer (Min)",    "Mehr Schlaf → mehr Energie?"),
        ("HRV (ms)",             "Höhere HRV → mehr Energie?"),
        ("REM (Min)",            "Mehr REM-Schlaf → mehr Energie?"),
        ("Tiefschlaf (Min)",     "Mehr Tiefschlaf → mehr Energie?"),
        ("Wachphasen (Min)",     "Mehr Wachphasen → weniger Energie?"),
        ("Schritte",             "Mehr Schritte Vortag → weniger Energie?"),
        ("Aktive Zeit (Min)",    "Mehr aktive Zeit → weniger Energie?"),
        ("Holy (Anzahl)",        "Mehr Holy Energy → mehr oder weniger Energie?"),
        ("Wasser: Gesamt (ml)",  "Mehr Wasser → mehr Energie?"),
        ("Schlafpuls (bpm)",     "Höherer Schlafpuls → weniger Energie?"),
    ]

					# Was passiert hier?
					# Für jeden Faktor werden nur die Tage genommen wo beide Werte vorhanden sind (z.B. Wasser nur bei neuen Logs). Dann 
					# berechnet stats.pearsonr die Korrelation. Bei weniger als 3 gemeinsamen Datenpunkten wird "Zu wenig Daten" 
					# angezeigt statt eines unsinnigen Wertes.

    # Energiewerte als Basis für alle Korrelationen
    energie_werte = [
        tag["Energiewert"] for tag in alle_daten
        if tag.get("Energiewert") is not None
    ]

    for zeile_idx, (faktor, frage) in enumerate(faktoren, start=start_zeile + 1):
        # Nur Tage nehmen wo beide Werte vorhanden sind
        paare = [
            (tag.get(faktor), tag.get("Energiewert"))
            for tag in alle_daten
            if tag.get(faktor) not in (None, "", "kein")
            and tag.get("Energiewert") is not None
        ]

        if len(paare) < 3:
            # Zu wenig Daten für eine sinnvolle Korrelation
            korr_wert  = "–"
            staerke    = "Zu wenig Daten"
            bedeutung  = frage
        else:
            x_werte = [float(str(p[0]).replace(",", ".")) for p in paare]
            y_werte = [float(p[1]) for p in paare]
            korr, _ = stats.pearsonr(x_werte, y_werte)
            korr_wert = round(korr, 3)

            # Stärke in Klartext übersetzen
            if abs(korr) >= 0.7:
                staerke = "Stark"
            elif abs(korr) >= 0.4:
                staerke = "Mittel"
            else:
                staerke = "Schwach"

            bedeutung = frage

					# Was passiert hier?
					# Jede Korrelationszeile wird geschrieben. Positive Korrelationen über 0.4 werden grün markiert, negative unter -0.4 werden 
					# rot markiert – so siehst du auf einen Blick was relevant ist.

        # Zeile schreiben
        zelle_f = ws3.cell(row=zeile_idx, column=1, value=faktor)
        zelle_k = ws3.cell(row=zeile_idx, column=2, value=korr_wert)
        zelle_s = ws3.cell(row=zeile_idx, column=3, value=staerke)
        zelle_b = ws3.cell(row=zeile_idx, column=4, value=bedeutung)

        for z in [zelle_f, zelle_k, zelle_s, zelle_b]:
            z.fill = PatternFill("solid", fgColor=FARBE_ZELLE)
            z.font = Font(color=FARBE_TEXT)

        # Farbe je nach Korrelationswert
        if isinstance(korr_wert, float):
            if korr_wert >= 0.4:
                zelle_k.font = Font(color=FARBE_GRUEN, bold=True)
            elif korr_wert <= -0.4:
                zelle_k.font = Font(color=FARBE_ROT, bold=True)

					# Was passiert hier?
					# Die fertige Excel-Datei wird gespeichert. In der Konsole erscheint eine Bestätigung mit der Anzahl der verarbeiteten Tage.

    # =====================
    # SPEICHERN
    # =====================
    wb.save(ausgabe_datei)
    print(f"✅ {len(alle_daten)} Einträge exportiert → {ausgabe_datei}")

					# Was passiert hier?
					# Pfade anpassen, Skript starten – fertig. Die .md ist die Quelldatei, die .xlsx wird neu erstellt (oder überschrieben falls 
					# sie schon existiert).

# --- START ---
md_pfad    = "/home/DEINNAME/pfad/zur/datei/energie_log.md"
excel_pfad = "/home/DEINNAME/Dokumente/energie_log.xlsx"

daten = parse_log(md_pfad)
in_excel_exportieren(daten, excel_pfad)
